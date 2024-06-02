import pandas as pd
import os
import argparse
import subprocess
import configparser
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
from tqdm import tqdm
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
import re
import time

def validate_data(data):
    valid_data = []
    for item in data:
        hostname, account, value = item
        hostname = re.sub(r'[^\x00-\x7F]+', '', hostname)
        account = re.sub(r'[^\x00-\x7F]+', '', account)
        value = re.sub(r'[^\x00-\x7F]+', '', value)
        valid_data.append([hostname, account, value])
    return valid_data

def process_secrets_files(directory):
    results = []
    for file in os.listdir(directory):
        if file.endswith('.secrets'):
            file_path = os.path.join(directory, file)
            if os.path.isfile(file_path) and os.access(file_path, os.R_OK):
                hostname = os.path.basename(file).replace('.secretsdump.secrets', '').lower()
                with open(file_path, 'r') as f:
                    for line in f:
                        if "SCM:{" in line or any(keyword in line for keyword in ['aes256', 'aes128', 'plain_password', 'des-cbc', 'dpapi', 'NL$KM', 'L$ASP.NET', 'L$_RasConn', 'aad3b435b51404eeaad3b435b51404ee', 'Security', 'RasDial', '| ', 'Version']):
                            continue
                        if ':' in line:
                            account, password = line.split(':', 1)
                            password = password.strip()
                            if len(password) > 50:
                                continue
                            results.append([hostname, account.strip().lower(), password])
    return validate_data(results)

def process_sam_files(directory):
    results = []
    for file in os.listdir(directory):
        if file.endswith('.sam'):
            file_path = os.path.join(directory, file)
            if os.path.isfile(file_path) and os.access(file_path, os.R_OK):
                hostname = os.path.basename(file).replace('.secretsdump.sam', '').lower()
                with open(file_path, 'r') as f:
                    for line in f:
                        if any(keyword in line for keyword in ['Default', 'Guest', 'WDAGUtility']):
                            continue
                        parts = line.strip().split(':')
                        if len(parts) >= 4:
                            account = parts[0].lower()
                            nt_hash = parts[3]
                            if nt_hash != '31d6cfe0d16ae931b73c59d7e0c089c0' and not account.startswith('_sc_gmsa_'):
                                results.append([hostname, account, nt_hash])
    return validate_data(results)

def get_pwned_label(tool):
    config_path = os.path.expanduser(f'~/.{tool}/{tool}.config')
    config = configparser.ConfigParser()
    if os.path.exists(config_path):
        config.read(config_path)
        if 'pwn3d_label' in config['DEFAULT']:
            return config['DEFAULT']['pwn3d_label']
    return 'Pwn3d'

def log_message(message, log_file):
    with open(log_file, 'a') as log:
        log.write(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - {message}\n")

def verify_local_admin_access(hostname, account, nt_hash, pwned_label, log_file):
    netexec_exists = subprocess.run(['which', 'netexec'], capture_output=True, text=True).returncode == 0
    if netexec_exists:
        command = f'netexec smb {hostname} -u {account} -H {nt_hash} --local-auth'
    else:
        print("\nNetExec doesn't exist, using CrackMapExec instead...\n")
        log_message("NetExec doesn't exist, using CrackMapExec instead...", log_file)
        command = f'crackmapexec smb {hostname} -u {account} -H {nt_hash} --local-auth'

    try:
        result = subprocess.run(command, shell=True, capture_output=True, text=True, timeout=60)
        log_message(f"Command: {command}\nOutput:\n{result.stdout}", log_file)
        if pwned_label in result.stdout:
            return (hostname, account, nt_hash)
    except subprocess.TimeoutExpired:
        error_message = f"Timeout expired for {command}"
        print(f"\n{error_message}\n")
        log_message(error_message, log_file)
    except Exception as e:
        error_message = f"Error verifying local admin access for {account}@{hostname}: {e}"
        print(f"\n{error_message}\n")
        log_message(error_message, log_file)
    
    return None

def apply_styles(sheet):
    header_fill = PatternFill(start_color='000080', end_color='000080', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    border_style = Side(border_style='thin', color='000000')
    border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.border = border

    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_letter].width = adjusted_width

def create_sanitized_file(df_secrets, df_sam, output_file):
    sanitized_output_file = output_file.replace(".xlsx", "_Sanitized.xlsx")

    with pd.ExcelWriter(sanitized_output_file, engine='openpyxl') as writer:
        df_secrets.drop(columns=['PASSWORD'], inplace=True)
        df_sam.drop(columns=['NT HASH'], inplace=True)
        df_sam = df_sam.sort_values(by=['ACCOUNT'])
        df_secrets.to_excel(writer, index=False, sheet_name='Service Account Cleartext Audit')
        df_sam.to_excel(writer, index=False, sheet_name='Local Admin Reuse Audit')

        workbook = writer.book
        secrets_sheet = writer.sheets['Service Account Cleartext Audit']
        sam_sheet = writer.sheets['Local Admin Reuse Audit']

        apply_styles(secrets_sheet)
        apply_styles(sam_sheet)

    print(f"\n[+] Sanitized file created: {os.path.abspath(sanitized_output_file)}\n")

def create_unverified_file(df_secrets, df_sam, output_file):
    unverified_output_file = output_file.replace(".xlsx", "_Unverified.xlsx")

    with pd.ExcelWriter(unverified_output_file, engine='openpyxl') as writer:
        df_secrets.to_excel(writer, index=False, sheet_name='Service Account Cleartext Audit')
        df_sam.to_excel(writer, index=False, sheet_name='Local Admin Reuse Audit')

        workbook = writer.book
        secrets_sheet = writer.sheets['Service Account Cleartext Audit']
        sam_sheet = writer.sheets['Local Admin Reuse Audit']

        apply_styles(secrets_sheet)
        apply_styles(sam_sheet)

    print(f"\n[+] Unverified file created: {os.path.abspath(unverified_output_file)}")

def write_to_excel(df_secrets, df_sam, output_file, log_file):
    try:
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df_secrets.to_excel(writer, index=False, sheet_name='Service Account Cleartext Audit')
            df_sam.to_excel(writer, index=False, sheet_name='Local Admin Reuse Audit')

            workbook = writer.book
            secrets_sheet = writer.sheets['Service Account Cleartext Audit']
            sam_sheet = writer.sheets['Local Admin Reuse Audit']

            apply_styles(secrets_sheet)
            apply_styles(sam_sheet)

        log_message(f"Successfully wrote data to {output_file}", log_file)
        print(f"\n[+] Success! ---> DumpInspector results can be found at {os.path.abspath(output_file)}")
    except Exception as e:
        log_message(f"Failed to write data to {output_file}: {e}", log_file)
        print(f"\n[!] Failed to write data to {output_file}: {e}\n")

def sanitize_value(value):
    if isinstance(value, str):
        return re.sub(r'[^\x00-\x7F]+', '', value)
    return value

def sanitize_dataframe(df):
    for column in df.columns:
        df[column] = df[column].apply(sanitize_value)
    return df

def main():
    parser = argparse.ArgumentParser(description="DumpInspector - A tool for auditing domain hosts' credential dumps.")
    parser.add_argument('-d', '--directory', help="<path-to-secretsdump-folder> containing the secretsdump output files.", required=True)
    parser.add_argument('-o', '--output', help="Output Excel file name (must end with .xlsx).")
    parser.add_argument('--no-verify', action='store_true', help="Skip the local admin verification step.")

    args = parser.parse_args()

    if not args.output:
        args.output = "DumpInspector_Results.xlsx"
    elif not args.output.endswith('.xlsx'):
        parser.error("Output file name must end with .xlsx")

    log_file = 'debug.log'
    with open(log_file, 'a') as log:
        log.write(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - [+] DumpInspector Log\n")

    print("\n[+] ... Auditing Secretsdump output...")
    log_message("Auditing Secretsdump output...", log_file)

    secrets_data = process_secrets_files(args.directory)
    df_secrets = pd.DataFrame(secrets_data, columns=['HOST', 'ACCOUNT', 'PASSWORD']).drop_duplicates()

    if df_secrets.empty:
        print("\nNote: No Service Account Credentials in Plaintext identified from Secretsdump output.\n")
        log_message("Note: No Service Account Credentials in Plaintext identified from Secretsdump output.", log_file)

    sam_data = process_sam_files(args.directory)
    df_sam = pd.DataFrame(sam_data, columns=['HOST', 'ACCOUNT', 'NT HASH'])

    if not args.no_verify:
        create_unverified_file(df_secrets, df_sam, args.output)
        try:
            verify_admin = input("\nAudit of Secretsdump data complete, would you like to confirm local admin results using NetExec? (Y/N): ").strip().lower()
        except EOFError:
            verify_admin = 'n'
        
        if verify_admin == 'y':
            print()
            pwned_label = get_pwned_label('nxc') if subprocess.run(['which', 'netexec'], capture_output=True, text=True).returncode == 0 else get_pwned_label('cme')
            
            verified_sam_data = []
            with tqdm(total=len(df_sam), desc="Verifying local admin access") as pbar:
                with ThreadPoolExecutor(max_workers=10) as executor:
                    futures = {executor.submit(verify_local_admin_access, row.HOST, row.ACCOUNT, row._2, pwned_label, log_file): row for row in df_sam.itertuples(index=False)}
                    for future in as_completed(futures):
                        result = future.result()
                        if result:
                            verified_sam_data.append(result)
                        pbar.update(1)
                        time.sleep(0.1)
            
            df_sam = pd.DataFrame(verified_sam_data, columns=['HOST', 'ACCOUNT', 'NT HASH'])
            if df_sam.empty:
                print("\nNote: No valid local admin re-use identified from Secretsdump output.")
                log_message("Note: No valid local admin re-use identified from Secretsdump output.", log_file)
        else:
            print("\n[!] Skipping local admin validation checks...\n")
            log_message("Skipping local admin validation checks...", log_file)
            return

    # Keep only duplicate accounts with the same name and NT hash
    df_sam = df_sam[df_sam.duplicated(subset=['ACCOUNT', 'NT HASH'], keep=False)]
    df_sam = df_sam.sort_values(by=['NT HASH', 'ACCOUNT'])

    df_secrets = sanitize_dataframe(df_secrets)
    df_sam = sanitize_dataframe(df_sam)

    write_to_excel(df_secrets, df_sam, args.output, log_file)

    try:
        create_sanitized = input("\nWould you like to create a sanitized version of the output file? (Y/N): ").strip().lower()
    except EOFError:
        create_sanitized = 'n'

    if create_sanitized == 'y':
        create_sanitized_file(df_secrets, df_sam, args.output)
        log_message("Sanitized version of the output file created.", log_file)

if __name__ == "__main__":
    main()
